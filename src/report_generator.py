"""
Report generator for filling the output template.
"""
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from typing import Optional, BinaryIO, Union
from io import BytesIO
from datetime import datetime

from .models import TransactionGroup, TransactionEntry, ProcessingResult, ReportSection, BANK_USD, BANK_VND
from config.mappings import (
    DEFAULT_OUTPUT_TEMPLATE,
    TEMPLATE_COL_VND,
    TEMPLATE_COL_USD,
    TEMPLATE_ROWS,
    NATURE_DISPLAY_MAP,
    INCOME_TYPE_MAP,
)


# Yellow highlight for manual review
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class ReportGenerator:
    """Generates filled report from processing results."""
    
    def __init__(self, template_source: Optional[Union[str, Path, BinaryIO]] = None):
        """
        Initialize with output template.
        
        Args:
            template_source: Path to template file or file-like object.
                            Uses default if not provided.
        """
        self.template_source = template_source or DEFAULT_OUTPUT_TEMPLATE
    
    def generate_report(self, result: ProcessingResult) -> bytes:
        """
        Generate filled report Excel file.
        
        Args:
            result: Processing result containing all calculated values
            
        Returns:
            Bytes of the filled Excel file
        """
        # Load template
        wb = load_workbook(self.template_source)
        ws = wb.active
        
        # Fill income section
        self._fill_income(ws, result)
        
        # Fill advance/settlement section
        self._fill_advance_settlement(ws, result)
        
        # Fill nature section
        self._fill_nature(ws, result)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    def _fill_income(self, ws, result: ProcessingResult) -> None:
        """Fill income section values with totals."""
        for bank_id, values in result.income.items():
            col = self._get_column(bank_id)
            
            # Contribution
            self._set_cell(ws, TEMPLATE_ROWS["contribution"], col, values.get("contribution", 0))
            
            # Fund transfer
            self._set_cell(ws, TEMPLATE_ROWS["fund_transfer"], col, values.get("fund_transfer", 0))
            
            # Fund transfer to ELC
            self._set_cell(ws, TEMPLATE_ROWS["fund_transfer_elc"], col, values.get("fund_transfer_elc", 0))
            
            # Interest
            self._set_cell(ws, TEMPLATE_ROWS["interest"], col, values.get("interest", 0))
            
            # PED
            self._set_cell(ws, TEMPLATE_ROWS["ped"], col, values.get("ped", 0))
            
            # Total for income section
            income_total = sum(values.values())
            if "income_total" in TEMPLATE_ROWS:
                self._set_cell(ws, TEMPLATE_ROWS["income_total"], col, income_total)
    
    def _fill_advance_settlement(self, ws, result: ProcessingResult) -> None:
        """Fill advance/settlement section values with totals."""
        for bank_id, values in result.advance_settlement.items():
            col = self._get_column(bank_id)
            
            # Advance
            self._set_cell(ws, TEMPLATE_ROWS["advance"], col, values.get("advance", 0))
            
            # Settlement
            self._set_cell(ws, TEMPLATE_ROWS["settlement"], col, values.get("settlement", 0))
            
            # Total for advance/settlement section
            total = sum(values.values())
            if "advance_settlement_total" in TEMPLATE_ROWS:
                self._set_cell(ws, TEMPLATE_ROWS["advance_settlement_total"], col, total)
    
    def _fill_nature(self, ws, result: ProcessingResult) -> None:
        """Fill expenditure by nature section values with totals."""
        for bank_id, values in result.nature_totals.items():
            col = self._get_column(bank_id)
            
            # Org
            self._set_cell(ws, TEMPLATE_ROWS["org"], col, values.get("org", 0))
            
            # Edu
            self._set_cell(ws, TEMPLATE_ROWS["edu"], col, values.get("edu", 0))
            
            # Oper
            self._set_cell(ws, TEMPLATE_ROWS["oper"], col, values.get("oper", 0))
            
            # Nutrition
            self._set_cell(ws, TEMPLATE_ROWS["nutrition"], col, values.get("nutrition", 0))
            
            # Edu Infra
            self._set_cell(ws, TEMPLATE_ROWS["edu_infra"], col, values.get("edu_infra", 0))
            
            # Total for expense section (excluding manual)
            expense_total = sum(v for k, v in values.items() if k != "manual")
            if "expense_total" in TEMPLATE_ROWS:
                self._set_cell(ws, TEMPLATE_ROWS["expense_total"], col, expense_total)
            
            # Manual
            self._set_cell(ws, TEMPLATE_ROWS["manual"], col, values.get("manual", 0))
    
    def _get_column(self, bank_id: str) -> int:
        """Get column index for bank type (1-based for openpyxl)."""
        if bank_id == BANK_VND:
            return TEMPLATE_COL_VND + 1  # Column D (4)
        return TEMPLATE_COL_USD + 1  # Column E (5)
    
    def _set_cell(self, ws, row: int, col: int, value: float) -> None:
        """Set cell value if non-zero."""
        if value != 0:
            ws.cell(row=row + 1, column=col, value=value)  # +1 because openpyxl is 1-based


def generate_marked_transactions(
    original_file: Union[str, Path, BinaryIO],
    all_groups: list[TransactionGroup],
    exchange_rates: dict[str, float],
    nature_mapper=None
) -> bytes:
    """
    Generate a processed transaction file with additional columns:
    - Report_Section (Income/Nature/Advance_Settlement)
    - Type (specific category within the section)
    - Amount_in_Respective_Currency
    - Is_Processed
    
    Args:
        original_file: Path or file-like object of original transaction file
        all_groups: List of all processed transaction groups
        exchange_rates: Dictionary of date string to exchange rate
        nature_mapper: NatureMapper instance for looking up nature categories
        
    Returns:
        Bytes of the enhanced Excel file
    """
    wb = load_workbook(original_file)
    ws = wb.active
    
    # Find the last column and add new headers
    max_col = ws.max_column
    
    # Add new column headers (find header row first)
    header_row = 1
    for row in range(1, min(10, ws.max_row + 1)):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and str(cell_val).lower() == "date":
            header_row = row
            break
    
    # New columns
    col_report_section = max_col + 1
    col_type = max_col + 2
    col_amount_currency = max_col + 3
    col_is_processed = max_col + 4
    col_exchange_rate = max_col + 5
    col_bank = max_col + 6
    
    # Set headers
    headers = [
        ("Report_Section", col_report_section),
        ("Type", col_type),
        ("Amount_in_Respective_Currency", col_amount_currency),
        ("Is_Processed", col_is_processed),
        ("Exchange_Rate", col_exchange_rate),
        ("Bank", col_bank),
    ]
    
    for header_text, col in headers:
        cell = ws.cell(row=header_row, column=col)
        cell.value = header_text
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    # Collect all row indices that need highlighting (manual review)
    highlight_rows: set[int] = set()
    
    # Process each group
    for group in all_groups:
        # Get exchange rate for this group's date
        date_key = group.date.strftime("%Y-%m-%d") if group.date else ""
        ex_rate = exchange_rates.get(date_key) if exchange_rates else None
        
        # Bank label
        bank_label = group.bank_identifier
        
        # Get section and type labels
        section_label, type_label = _get_section_type_labels(group)
        
        # Mark manual review rows
        if group.assigned_section == ReportSection.MANUAL or any(e.is_manual_trigger for e in group.entries):
            highlight_rows.add(group.original_row_index + 1)
            for entry in group.entries:
                highlight_rows.add(entry.original_row_index + 1)
        
        # Fill header row
        excel_row = group.original_row_index + 1
        if excel_row > header_row:
            ws.cell(row=excel_row, column=col_report_section).value = section_label
            ws.cell(row=excel_row, column=col_type).value = type_label
            ws.cell(row=excel_row, column=col_is_processed).value = "Yes" if group.is_processed else "No"
            ws.cell(row=excel_row, column=col_bank).value = bank_label
            
            # Exchange rate (only for USD bank)
            if group.bank_identifier == BANK_USD and ex_rate:
                ws.cell(row=excel_row, column=col_exchange_rate).value = ex_rate
            
            # Converted amount
            if group.bank_amount:
                if group.bank_identifier == BANK_USD and ex_rate:
                    converted = group.bank_amount / ex_rate
                    ws.cell(row=excel_row, column=col_amount_currency).value = converted
                else:
                    ws.cell(row=excel_row, column=col_amount_currency).value = group.bank_amount
        
        # Fill entry rows
        for entry in group.entries:
            excel_row = entry.original_row_index + 1
            
            if excel_row <= header_row:
                continue
            
            # Determine entry type based on group's section
            # For INCOME/ADVANCE_SETTLEMENT: always inherit group's type_label
            # For NATURE/MANUAL: use entry's determined nature_type (never "manual")
            entry_type = type_label  # Default to group's type
            
            if group.assigned_section in (ReportSection.NATURE, ReportSection.MANUAL):
                # Use entry's nature_type if available and not "manual"
                if entry.nature_type and entry.nature_type != "manual":
                    entry_type = NATURE_DISPLAY_MAP.get(entry.nature_type, entry.nature_type)
            
            ws.cell(row=excel_row, column=col_report_section).value = section_label
            ws.cell(row=excel_row, column=col_type).value = entry_type
            ws.cell(row=excel_row, column=col_is_processed).value = "Yes" if group.is_processed else "No"
            ws.cell(row=excel_row, column=col_bank).value = bank_label
            
            if group.bank_identifier == BANK_USD and ex_rate:
                ws.cell(row=excel_row, column=col_exchange_rate).value = ex_rate
            
            # Converted amount for entry
            if entry.amount:
                if group.bank_identifier == BANK_USD and ex_rate:
                    converted = entry.amount / ex_rate
                    ws.cell(row=excel_row, column=col_amount_currency).value = converted
                else:
                    ws.cell(row=excel_row, column=col_amount_currency).value = entry.amount
    
    # Apply highlighting for manual review rows
    for row_idx in highlight_rows:
        for col in range(1, col_bank + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = HIGHLIGHT_FILL
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _get_section_type_labels(group: TransactionGroup) -> tuple[str, str]:
    """
    Get section and type labels for a transaction group.
    
    Returns:
        Tuple of (section_label, type_label)
    """
    section = group.assigned_section
    
    if section == ReportSection.INCOME:
        section_label = "Income"
        # Determine which income type
        if group.is_deposit() and group.any_name_contains("onesky"):
            type_label = "Contribution"
        elif group.any_memo_contains("transfer") and group.any_memo_contains("elc"):
            type_label = "Fund transfer to ELC"
        elif group.any_memo_contains("transfer") and not group.any_memo_contains("elc"):
            type_label = "Fund transfer to USD"
        elif group.is_deposit() and group.any_memo_contains("interest"):
            type_label = "Interest"
        elif group.is_deposit() and group.any_memo_contains("ped"):
            type_label = "PED"
        else:
            type_label = ""
        return section_label, type_label
    
    elif section == ReportSection.ADVANCE_SETTLEMENT:
        section_label = "Advance_Settlement"
        if group.any_memo_contains("advance") and not group.any_memo_contains("settlement"):
            type_label = "Advance by cash"
        elif group.any_memo_contains("settlement"):
            type_label = "Settlement"
        else:
            type_label = ""
        return section_label, type_label
    
    elif section == ReportSection.NATURE:
        section_label = "Nature"
        # Get nature from first active entry
        active = group.active_entries
        if active and active[0].nature_type:
            type_label = NATURE_DISPLAY_MAP.get(active[0].nature_type, active[0].nature_type)
        else:
            type_label = ""
        return section_label, type_label
    
    elif section == ReportSection.MANUAL:
        section_label = "Advance_Settlement"  # Manual groups are typically advance/settlement
        # Find the first entry with a determined nature_type
        active = group.active_entries
        type_label = ""
        for entry in active:
            if entry.nature_type and entry.nature_type != "manual":
                type_label = NATURE_DISPLAY_MAP.get(entry.nature_type, entry.nature_type)
                break
        # If no nature found, try to determine from memo
        if not type_label:
            if group.any_memo_contains("advance") and not group.any_memo_contains("settlement"):
                type_label = "Advance by cash"
            elif group.any_memo_contains("settlement"):
                type_label = "Settlement"
        return section_label, type_label
    
    # Default
    return "", ""
